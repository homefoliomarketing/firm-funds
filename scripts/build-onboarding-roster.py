"""Build the Firm Funds Brokerage Agent Onboarding Roster (.xlsx).

Pre-filled with the live Century 21 Choice Realty roster, but designed to be
reused as the onboarding template for any new brokerage going forward.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

LOGO = r"C:\Users\randi\Dev\firm-funds\public\brand\icon-white.png"

def place_logo(ws, w_px, h_px, x_px, y_px):
    img = XLImage(LOGO)
    img.width = w_px; img.height = h_px
    marker = AnchorMarker(col=0, colOff=pixels_to_EMU(x_px), row=0, rowOff=pixels_to_EMU(y_px))
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(w_px), pixels_to_EMU(h_px)))
    ws.add_image(img)

# ---- Brand palette (from lib/email.ts) -------------------------------------
GREEN       = "5FA873"   # primary brand green
DEEP_GREEN  = "0D2818"   # near-black green for the title bar
MID_GREEN   = "1A4D2E"
GOLD        = "D4A04A"   # accent
TINT        = "F2F8F4"   # very light green, zebra striping
TINT_HEAD   = "EAF3EE"   # section header fill
INPUT_YEL   = "FFF8E1"   # soft yellow = "fill me in"
GREY_TXT    = "737373"
GREY_LINE   = "DCE3DD"
WHITE       = "FFFFFF"
DARK_TXT    = "1E1E1E"

FONT = "Arial"

def f(size=10, bold=False, color=DARK_TXT, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def fill(hexcol):
    return PatternFill("solid", fgColor=hexcol)

thin = Side(style="thin", color=GREY_LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom_only = Border(bottom=Side(style="thin", color=GREEN))

LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_W = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ---- Roster data (live from Supabase, Choice Advances brokerage) -----------
# Junk test emails on real agents (invalid domains) were blanked; valid-looking
# emails/phones kept. Phones normalized to ###-###-####.
AGENTS = [
    ("Marc",    "Beaudette",     "",                         "705-542-1016"),
    ("Mike",    "Bell",          "",                         ""),
    ("Larry",   "Berto",         "",                         ""),
    ("Lance",   "Brizard",       "",                         ""),
    ("Lee",     "Bryar",         "",                         ""),
    ("James",   "Caicco",        "james.caicco@century21.ca","705-941-8978"),
    ("Thomas",  "Campana",       "",                         ""),
    ("Carrie",  "Cerenzie",      "",                         ""),
    ("Stacey",  "Cleave",        "",                         ""),
    ("Sylvie",  "D'Ettore",      "",                         ""),
    ("Ryan",    "Dodd",          "",                         "705-542-1016"),
    ("Dave",    "Edgar",         "",                         ""),
    ("Terry",   "Esposti",       "",                         ""),
    ("Bill",    "Fraser",        "",                         ""),
    ("Mike",    "Gallivan",      "",                         ""),
    ("Carlo",   "Gervasi",       "",                         ""),
    ("Nikki",   "Gilbert",       "",                         ""),
    ("Dallas",  "Glawson",       "",                         ""),
    ("Anthony", "Greco",         "",                         ""),
    ("Susan",   "Hawke",         "",                         ""),
    ("Sandra",  "Hunter-Dunn",   "",                         ""),
    ("Joey",    "Iachetta",      "",                         ""),
    ("Bud",     "Jones",         "bud.jones@century21.ca",   "705-542-1016"),
    ("Tricia",  "Kent",          "",                         ""),
    ("Joanne",  "Kovich",        "",                         ""),
    ("Avery",   "Marcoux",       "",                         ""),
    ("Brianne", "McGill",        "",                         ""),
    ("Jodie",   "McNabb",        "",                         ""),
    ("Elinor",  "Mick",          "",                         ""),
    ("Bill",    "Montague",      "",                         ""),
    ("Frank",   "Naccarato",     "",                         ""),
    ("Dan",     "Nogalo",        "",                         ""),
    ("Tammy",   "Ovey",          "",                         ""),
    ("Kristen", "Peltsch",       "",                         ""),
    ("Andrew",  "Raplenovic",    "",                         ""),
    ("Tiffany", "Rogers",        "",                         ""),
    ("Mike",    "Rotermann",     "",                         ""),
    ("Kyle",    "Scali",         "",                         ""),
    ("Patricia","Shell",         "",                         ""),
    ("Jason",   "Sproule",       "",                         ""),
    ("Gary",    "Trembinski",    "",                         ""),
    ("Glen",    "Trembinski",    "",                         ""),
    ("Ken",     "Vandaele",      "",                         "705-542-1016"),
    ("Bill",    "Vanderleest",   "",                         ""),
    ("Katie",   "Watkins",       "",                         ""),
    ("Liz",     "Willson",       "",                         ""),
    ("Andrea",  "Zavitz",        "",                         ""),
]

wb = Workbook()
wb.calculation.fullCalcOnLoad = True  # force recalc of COUNTA on open

# ===========================================================================
#  SHEET 1 — Agent Roster
# ===========================================================================
ws = wb.active
ws.title = "Agent Roster"
ws.sheet_view.showGridLines = False

COLS = ["#", "First Name", "Last Name", "Email", "Phone",
        "RECO Registration #", "Status", "Notes"]
WIDTHS = [5, 15, 18, 30, 15, 21, 12, 26]
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
LAST_COL = get_column_letter(len(COLS))  # 'H'

# ---- Title band -----------------------------------------------------------
for col in range(1, len(COLS) + 1):
    ws.cell(row=1, column=col).fill = fill(DEEP_GREEN)
    ws.cell(row=2, column=col).fill = fill(DEEP_GREEN)
ws.merge_cells(f"B1:{LAST_COL}1")
ws.merge_cells(f"B2:{LAST_COL}2")
c = ws["B1"]
c.value = "FIRM FUNDS   ·   COMMISSION ADVANCES"
c.font = f(10, bold=True, color=GREEN)
c.alignment = Alignment(horizontal="left", vertical="bottom", indent=5)
c2 = ws["B2"]
c2.value = "Brokerage Agent Onboarding Roster"
c2.font = f(22, bold=True, color=WHITE)
c2.alignment = Alignment(horizontal="left", vertical="top", indent=5)
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 38
place_logo(ws, 51, 52, 8, 13)
# gold accent rule
ws.merge_cells(f"A3:{LAST_COL}3")
for col in range(1, len(COLS) + 1):
    ws.cell(row=3, column=col).fill = fill(GOLD)
ws.row_dimensions[3].height = 5

# ---- Brokerage details block ----------------------------------------------
ws.row_dimensions[4].height = 6  # spacer
ws.merge_cells(f"A5:{LAST_COL}5")
sh = ws["A5"]
sh.value = "BROKERAGE DETAILS"
sh.font = f(10, bold=True, color=MID_GREEN)
sh.fill = fill(TINT_HEAD)
sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[5].height = 22

def label(cell, text):
    cell.value = text
    cell.font = f(8, bold=True, color=GREY_TXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def input_cell(anchor_col, row, span_to, value="", formula=False):
    """Merge anchor..span_to on `row`, style as a fill-in (or computed) cell."""
    a = get_column_letter(anchor_col)
    b = get_column_letter(span_to)
    ws.merge_cells(f"{a}{row}:{b}{row}")
    cell = ws[f"{a}{row}"]
    cell.value = value
    cell.font = f(10, bold=True, color=DARK_TXT if formula else "0000CC")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = fill(WHITE if formula else INPUT_YEL)
    cell.border = bottom_only
    return cell

# Two columns of label / value pairs across A-D (left) and E-H (right)
# Left labels in A, values B:D ; Right labels in E, values F:H
detail_rows = [
    ("Brokerage Name", "Century 21 Choice Realty", "Date Submitted", ""),
    ("Primary Admin Contact", "", "Admin Email", ""),
    ("Admin Phone", "", "Office Location", ""),
]
def detail_label(col_from, col_to, row, text):
    a = get_column_letter(col_from); b = get_column_letter(col_to)
    ws.merge_cells(f"{a}{row}:{b}{row}")
    label(ws[f"{a}{row}"], text)

r = 6
for left_lbl, left_val, right_lbl, right_val in detail_rows:
    ws.row_dimensions[r].height = 20
    detail_label(1, 2, r, left_lbl)     # label A:B
    input_cell(3, r, 4, left_val)       # value C:D
    detail_label(5, 6, r, right_lbl)    # label E:F
    input_cell(7, r, 8, right_val)      # value G:H
    r += 1
# Total agents (computed) row
ws.row_dimensions[r].height = 20
detail_label(1, 2, r, "Total Agents")
TABLE_HEADER_ROW = r + 3
first_data = TABLE_HEADER_ROW + 1
last_data = first_data + len(AGENTS) - 1
cnt = input_cell(3, r, 4, f"=COUNTA(C{first_data}:C{last_data})", formula=True)
detail_label(5, 6, r, "Onboarding Status")
input_cell(7, r, 8, "")
r += 1
ws.row_dimensions[r].height = 8  # spacer

# ---- Roster section header ------------------------------------------------
sec_row = r + 1
ws.merge_cells(f"A{sec_row}:{LAST_COL}{sec_row}")
sh2 = ws.cell(row=sec_row, column=1)
sh2.value = "AGENT ROSTER"
sh2.font = f(10, bold=True, color=MID_GREEN)
sh2.fill = fill(TINT_HEAD)
sh2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[sec_row].height = 22

assert sec_row + 1 == TABLE_HEADER_ROW, (sec_row, TABLE_HEADER_ROW)

# ---- Table header ----------------------------------------------------------
hr = TABLE_HEADER_ROW
ws.row_dimensions[hr].height = 26
for i, head in enumerate(COLS, start=1):
    cell = ws.cell(row=hr, column=i)
    cell.value = head
    cell.font = f(10, bold=True, color=WHITE)
    cell.fill = fill(GREEN)
    cell.alignment = CENTER if head in ("#", "Status") else Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = box

# ---- Data rows -------------------------------------------------------------
for idx, (first, last, email, phone) in enumerate(AGENTS):
    row = first_data + idx
    ws.row_dimensions[row].height = 19
    zebra = TINT if idx % 2 else WHITE
    values = [idx + 1, first, last, email, phone, "", "Active", ""]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.fill = fill(zebra)
        cell.border = box
        if col == 1:
            cell.font = f(9, color=GREY_TXT); cell.alignment = CENTER
        elif col == 7:  # Status
            cell.font = f(9, bold=True, color=GREEN); cell.alignment = CENTER
        else:
            cell.font = f(10, color=DARK_TXT)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ---- Status dropdown -------------------------------------------------------
dv = DataValidation(type="list", formula1='"Active,Inactive,Pending,Invited"', allow_blank=True)
dv.error = "Choose: Active, Inactive, Pending, or Invited"
dv.errorTitle = "Invalid status"
ws.add_data_validation(dv)
dv.add(f"G{first_data}:G{last_data}")

# ---- Freeze, print setup ---------------------------------------------------
ws.freeze_panes = f"A{first_data}"
ws.print_title_rows = f"{hr}:{hr}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5

# ===========================================================================
#  SHEET 2 — How to Use
# ===========================================================================
ws2 = wb.create_sheet("How to Use")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 96

for rr in (1, 2):
    ws2.cell(row=rr, column=1).fill = fill(DEEP_GREEN)
    ws2.cell(row=rr, column=2).fill = fill(DEEP_GREEN)
t1 = ws2["B1"]; t1.value = "FIRM FUNDS   ·   COMMISSION ADVANCES"
t1.font = f(10, bold=True, color=GREEN)
t1.alignment = Alignment(horizontal="left", vertical="bottom", indent=7)
t2 = ws2["B2"]; t2.value = "How to Use This Onboarding Roster"
t2.font = f(20, bold=True, color=WHITE)
t2.alignment = Alignment(horizontal="left", vertical="top", indent=7)
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 36
place_logo(ws2, 51, 52, 8, 13)
ws2.merge_cells("A3:B3")
for col in (1, 2):
    ws2.cell(row=3, column=col).fill = fill(GOLD)
ws2.row_dimensions[3].height = 5

def block(row, text, *, head=False, bullet=False, note=False):
    ws2.row_dimensions[row].height = 28 if head else 22
    cell = ws2.cell(row=row, column=2)
    cell.value = text
    if head:
        cell.font = f(12, bold=True, color=MID_GREEN)
    elif note:
        cell.font = f(9, italic=True, color=GREY_TXT)
    else:
        cell.font = f(10, color=DARK_TXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

rows2 = [
    (5,  "What this is", dict(head=True)),
    (6,  "A reusable roster Firm Funds uses to onboard every agent at a partner brokerage. It is pre-filled with the Century 21 Choice Realty roster as a working example.", {}),
    (7,  "", {}),
    (8,  "Steps", dict(head=True)),
    (9,  "1.  Fill in the yellow Brokerage Details fields at the top of the Agent Roster tab.", {}),
    (10, "2.  List every agent: first name and last name are required; email, phone, and RECO registration number are strongly preferred.", {}),
    (11, "3.  Set each agent's Status from the dropdown (Active, Inactive, Pending, or Invited). Default is Active.", {}),
    (12, "4.  The Total Agents count at the top updates automatically as you add or remove rows.", {}),
    (13, "5.  Send the completed file back to Firm Funds at homefoliomarketing@gmail.com to start onboarding.", {}),
    (14, "", {}),
    (15, "Field guide", dict(head=True)),
    (16, "Email / Phone  —  the agent's direct contact. Used for advance notifications and KYC.", {}),
    (17, "RECO Registration #  —  the agent's Real Estate Council of Ontario registration number.", {}),
    (18, "Status  —  Active means currently licensed and eligible; Invited means an invite was sent but not yet accepted.", {}),
    (19, "", {}),
    (20, "Legend:  yellow cells = fill in   |   blue text = your input   |   green = totals are calculated automatically.", dict(note=True)),
]
for row, text, kw in rows2:
    block(row, text, **kw)

ws2.page_setup.orientation = "portrait"
ws2.page_margins.left = ws2.page_margins.right = 0.6

OUT = r"C:\Users\randi\Dev\firm-funds\Firm-Funds-Agent-Onboarding-Roster.xlsx"
wb.save(OUT)
print("Saved:", OUT)
print("Agents:", len(AGENTS), "| header row:", hr, "| data rows:", first_data, "-", last_data)
